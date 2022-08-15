import requests
from bs4 import BeautifulSoup
import openpyxl

if __name__ == '__main__':
    # 엑셀 파일 만들기
    wb = openpyxl.Workbook()
    # 엑셀 워크시트 만들기
    ws = wb.create_sheet("증권 데이터")

    # 종목 코드 리스트
    codes = {
        "삼성전자": "005930",
        "SK하이닉스": "000660",
        "카카오": "035720"
    }

    for company, code in codes.items():
        url = f"https://finance.naver.com/item/sise.naver?code={code}"

        response = requests.get(url)
        html = response.text

        soup = BeautifulSoup(markup=html, features="html.parser")
        current_price = soup.select_one("#_nowVal")
        current_price_value = int(current_price.text.replace(",", ""))
        print(f"종목: {company}, 현재가: {current_price_value}")
        # 3) 셀에 데이터 추가하기
        ws["A1"]
