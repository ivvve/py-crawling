import openpyxl

# 1) 엑셀 만들기
wb = openpyxl.Workbook()
# 2) 엑셀 워크시트 만들기
ws = wb.create_sheet("참가자")
# 3) 셀에 데이터 추가하기
ws["A1"] = "참가번호"
ws["B1"] = "성명"
ws["A2"] = 1
ws["B2"] = "오일남"
# 4) 엑셀 파일 저장하기
wb.save("참가자.xlsx")

# 5) 엑셀 불러오기
wb = openpyxl.load_workbook("참가자.xlsx")
# 6) 엑셀 시트 선택
ws = wb["참가자"]
# 7) 데이터 수정하기
ws["A3"] = 456
ws["B3"] = "성기훈"
# 8) 엑셀 저장하기
wb.save("updated_참가자.xlsx")
